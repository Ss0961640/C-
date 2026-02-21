# -*- coding: utf-8 -*-
import os
import json
import numpy as np

import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader
from torchvision import datasets, transforms, models

from tqdm import tqdm

import matplotlib.pyplot as plt
from sklearn.metrics import (
    accuracy_score,
    precision_recall_fscore_support,
    classification_report,
    confusion_matrix,
)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import onnx
import onnx.checker


# ================= 絕對路徑設定 =================
TRAIN_DIR = r"C:\Users\Raymond_yang\Desktop\Autonomous_Ship\ships_dataset\Various_categories_8\train"
VAL_DIR   = r"C:\Users\Raymond_yang\Desktop\Autonomous_Ship\ships_dataset\Various_categories_8\valid"
TEST_DIR  = r"C:\Users\Raymond_yang\Desktop\Autonomous_Ship\ships_dataset\Various_categories_8\test"
OUT_DIR   = r"C:\Users\Raymond_yang\Desktop\Autonomous_Ship\ships_dataset\result\Various_categories_8_50"

# ===================== 參數 ====================
EPOCHS = 50

# EfficientNet-B7 很吃顯存：若 OOM -> 改成 4 或 2
BATCH_SIZE = 16

LR = 1e-4
WEIGHT_DECAY = 1e-4
NUM_WORKERS = 2
PATIENCE = 7

# ==========================================================
# 指定要跑的模型清單
# ==========================================================
MODEL_SPECS = [
    {"name": "shallow_cnn",         "img_size": 224},
    {"name": "vgg16",               "img_size": 224},
    {"name": "resnet50",            "img_size": 224},
    {"name": "resnet101",           "img_size": 224},
    {"name": "densenet121",         "img_size": 224},
    {"name": "inception_v3",        "img_size": 299},
    {"name": "mobilenet_v3_large",  "img_size": 224},
    {"name": "efficientnet_b7",     "img_size": 224},  # 若要用原始建議可改 600，但顯存需求會大增
]


# ===================== 工具函式 =====================
def compute_class_weights(imagefolder_dataset: datasets.ImageFolder):
    targets = np.array(imagefolder_dataset.targets)
    num_classes = len(imagefolder_dataset.classes)
    counts = np.bincount(targets, minlength=num_classes).astype(np.float64)
    total = counts.sum()
    weights = total / (num_classes * (counts + 1e-12))
    return torch.tensor(weights, dtype=torch.float32), counts


def build_transforms(img_size: int):
    mean, std = (0.485, 0.456, 0.406), (0.229, 0.224, 0.225)

    train_tf = transforms.Compose([
        transforms.Resize((img_size, img_size)),
        transforms.RandomHorizontalFlip(p=0.5),
        transforms.RandomRotation(10),
        transforms.ColorJitter(brightness=0.15, contrast=0.15, saturation=0.10),
        transforms.ToTensor(),
        transforms.Normalize(mean, std),
    ])

    val_tf = transforms.Compose([
        transforms.Resize((img_size, img_size)),
        transforms.ToTensor(),
        transforms.Normalize(mean, std),
    ])

    return train_tf, val_tf, mean, std


def _replace_last_linear_in_sequential(seq: nn.Sequential, num_classes: int):
    for i in range(len(seq) - 1, -1, -1):
        if isinstance(seq[i], nn.Linear):
            in_f = seq[i].in_features
            seq[i] = nn.Linear(in_f, num_classes)
            return
    raise ValueError("找不到可替換的 Linear 層（classifier 裡沒有 Linear）。")


def build_model(model_name: str, num_classes: int):
    if model_name == "resnet50":
        model = models.resnet50(weights=models.ResNet50_Weights.DEFAULT)
        model.fc = nn.Linear(model.fc.in_features, num_classes)
        return model, "resnet50"

    elif model_name == "resnet101":
        model = models.resnet101(weights=models.ResNet101_Weights.DEFAULT)
        model.fc = nn.Linear(model.fc.in_features, num_classes)
        return model, "resnet101"

    elif model_name == "densenet121":
        model = models.densenet121(weights=models.DenseNet121_Weights.DEFAULT)
        model.classifier = nn.Linear(model.classifier.in_features, num_classes)
        return model, "densenet121"

    elif model_name == "vgg16":
        model = models.vgg16(weights=models.VGG16_Weights.DEFAULT)
        _replace_last_linear_in_sequential(model.classifier, num_classes)
        return model, "vgg16"

    elif model_name == "mobilenet_v3_large":
        model = models.mobilenet_v3_large(weights=models.MobileNet_V3_Large_Weights.DEFAULT)
        _replace_last_linear_in_sequential(model.classifier, num_classes)
        return model, "mobilenet_v3_large"

    elif model_name == "efficientnet_b7":
        model = models.efficientnet_b7(weights=models.EfficientNet_B7_Weights.DEFAULT)
        _replace_last_linear_in_sequential(model.classifier, num_classes)
        return model, "efficientnet_b7"

    elif model_name == "inception_v3":
        model = models.inception_v3(weights=models.Inception_V3_Weights.DEFAULT, aux_logits=True)
        model.fc = nn.Linear(model.fc.in_features, num_classes)
        if model.AuxLogits is not None:
            model.AuxLogits.fc = nn.Linear(model.AuxLogits.fc.in_features, num_classes)
        return model, "inception_v3"

    elif model_name == "shallow_cnn":
        class ShallowCNN(nn.Module):
            def __init__(self, n_classes: int):
                super().__init__()
                self.features = nn.Sequential(
                    nn.Conv2d(3, 32, kernel_size=3, padding=1),
                    nn.BatchNorm2d(32),
                    nn.ReLU(inplace=True),
                    nn.MaxPool2d(2),

                    nn.Conv2d(32, 64, kernel_size=3, padding=1),
                    nn.BatchNorm2d(64),
                    nn.ReLU(inplace=True),
                    nn.MaxPool2d(2),

                    nn.Conv2d(64, 128, kernel_size=3, padding=1),
                    nn.BatchNorm2d(128),
                    nn.ReLU(inplace=True),
                    nn.MaxPool2d(2),
                )
                self.pool = nn.AdaptiveAvgPool2d((1, 1))
                self.classifier = nn.Sequential(
                    nn.Flatten(),
                    nn.Dropout(0.3),
                    nn.Linear(128, n_classes)
                )

            def forward(self, x):
                x = self.features(x)
                x = self.pool(x)
                x = self.classifier(x)
                return x

        return ShallowCNN(num_classes), "shallow_cnn"

    else:
        raise ValueError(f"Unknown model_name: {model_name}")


def forward_logits(model, x, model_name: str, use_amp: bool):
    """
    統一處理 forward 輸出
    - inception_v3 train: InceptionOutputs(logits, aux_logits)
    - others: logits
    """
    with torch.amp.autocast('cuda', enabled=use_amp):
        out = model(x)
        if model_name == "inception_v3":
            if isinstance(out, tuple) or hasattr(out, "logits"):
                logits = out.logits if hasattr(out, "logits") else out[0]
                aux = out.aux_logits if hasattr(out, "aux_logits") else (out[1] if len(out) > 1 else None)
                return logits, aux
        return out, None


@torch.no_grad()
def evaluate(model, loader, device, criterion, model_name: str):
    model.eval()
    total, correct = 0, 0
    running_loss = 0.0

    for x, y in tqdm(loader, desc="Eval", leave=False):
        x = x.to(device, non_blocking=True)
        y = y.to(device, non_blocking=True)

        logits, _ = forward_logits(model, x, model_name, use_amp=False)
        loss = criterion(logits.float(), y)

        running_loss += loss.item() * x.size(0)
        pred = logits.argmax(1)
        correct += (pred == y).sum().item()
        total += y.size(0)

    val_loss = running_loss / max(total, 1)
    val_acc = correct / max(total, 1)
    return val_loss, val_acc


@torch.no_grad()
def predict_with_paths(model, imagefolder_ds: datasets.ImageFolder, loader, device, model_name: str):
    model.eval()
    y_true, y_pred, confs = [], [], []

    for x, y in tqdm(loader, desc="Predict", leave=False):
        x = x.to(device, non_blocking=True)
        logits, _ = forward_logits(model, x, model_name, use_amp=False)
        probs = torch.softmax(logits, dim=1)
        p, pred = probs.max(dim=1)

        y_true.extend(y.numpy().tolist())
        y_pred.extend(pred.cpu().numpy().tolist())
        confs.extend(p.cpu().numpy().tolist())

    paths = [s[0] for s in imagefolder_ds.samples]  # shuffle=False 才能對齊
    return y_true, y_pred, confs, paths


def plot_curves_fixed_epochs(train_losses, val_losses, train_accs, val_accs, out_path, total_epochs: int):
    """
    X 軸固定顯示 1..total_epochs（early stop 也會補齊到 EPOCHS）
    """
    def pad_to_len(arr, L):
        arr = list(arr)
        if len(arr) < L:
            arr += [np.nan] * (L - len(arr))
        return arr[:L]

    tl = pad_to_len(train_losses, total_epochs)
    vl = pad_to_len(val_losses, total_epochs)
    ta = pad_to_len(train_accs, total_epochs)
    va = pad_to_len(val_accs, total_epochs)

    epochs = range(1, total_epochs + 1)

    plt.figure(figsize=(10, 4))

    plt.subplot(1, 2, 1)
    plt.plot(epochs, tl, label="train_loss")
    plt.plot(epochs, vl, label="val_loss")
    plt.xlabel("Epoch")
    plt.ylabel("Loss")
    plt.title("Loss Curve")
    plt.legend()

    plt.subplot(1, 2, 2)
    plt.plot(epochs, ta, label="train_acc")
    plt.plot(epochs, va, label="val_acc")
    plt.xlabel("Epoch")
    plt.ylabel("Accuracy")
    plt.title("Accuracy Curve")
    plt.legend()

    plt.tight_layout()
    plt.savefig(out_path, dpi=200)
    plt.close()


def plot_confusion_matrix(cm, class_names, out_path_png):
    plt.figure(figsize=(8, 7))
    plt.imshow(cm, interpolation="nearest", cmap=plt.cm.Blues)
    plt.title("Confusion Matrix")
    plt.colorbar()

    tick_marks = np.arange(len(class_names))
    plt.xticks(tick_marks, class_names, rotation=45, ha="right")
    plt.yticks(tick_marks, class_names)

    thresh = cm.max() / 2.0 if cm.max() > 0 else 0.0
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            plt.text(
                j, i, format(cm[i, j], "d"),
                ha="center", va="center",
                color="white" if cm[i, j] > thresh else "black"
            )

    plt.ylabel("True label")
    plt.xlabel("Predicted label")
    plt.tight_layout()
    plt.savefig(out_path_png, dpi=200)
    plt.close()


# ===================== ONNX 匯出（修正版，提升成功率） =====================
class ONNXWrapper(nn.Module):
    """
    讓輸出永遠只回 logits，避免 inception_v3 可能回 InceptionOutputs / tuple。
    """
    def __init__(self, model: nn.Module, model_name: str):
        super().__init__()
        self.model = model
        self.model_name = model_name

    def forward(self, x):
        out = self.model(x)
        if self.model_name == "inception_v3":
            if hasattr(out, "logits"):
                return out.logits
            if isinstance(out, (tuple, list)):
                return out[0]
        return out


def export_onnx_from_best_pt(best_pt_path: str, model_name: str, img_size: int, num_classes: int, onnx_path: str):
    ckpt = torch.load(best_pt_path, map_location="cpu")

    model, _ = build_model(model_name, num_classes)
    model.load_state_dict(ckpt["model_state"])
    model.eval()
    model = model.to("cpu").float()

    wrapped = ONNXWrapper(model, model_name).eval()
    dummy = torch.randn(1, 3, img_size, img_size, device="cpu", dtype=torch.float32)

    torch.onnx.export(
        wrapped,
        dummy,
        onnx_path,
        export_params=True,
        opset_version=13,
        do_constant_folding=True,
        input_names=["input"],
        output_names=["logits"],
        dynamic_axes={"input": {0: "batch"}, "logits": {0: "batch"}},
    )

    # ✅ 匯出後檢查檔案存在
    if (not os.path.exists(onnx_path)) or (os.path.getsize(onnx_path) == 0):
        raise RuntimeError(f"ONNX 匯出失敗：檔案不存在或大小為 0 -> {onnx_path}")

    # ✅ 用 onnx.checker 驗證模型結構是否正確
    onnx_model = onnx.load(onnx_path)
    onnx.checker.check_model(onnx_model)


def write_metrics_excel(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "metrics"

    headers = [
        "model",
        "img_size",
        "epochs_setting",
        "epochs_ran",
        "best_val_acc",
        "test_accuracy",
        "precision_weighted",
        "recall_weighted",
        "f1_weighted",
        "precision_macro",
        "recall_macro",
        "f1_macro",
        "pt_path",
        "onnx_path",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            r.get("model"),
            r.get("img_size"),
            r.get("epochs_setting"),
            r.get("epochs_ran"),
            r.get("best_val_acc"),
            r.get("test_accuracy"),
            r.get("precision_weighted"),
            r.get("recall_weighted"),
            r.get("f1_weighted"),
            r.get("precision_macro"),
            r.get("recall_macro"),
            r.get("f1_macro"),
            r.get("pt_path"),
            r.get("onnx_path"),
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    wb.save(out_path)


def run_one_model(model_name: str, img_size: int, device):
    model_out_dir = os.path.join(OUT_DIR, model_name)
    os.makedirs(model_out_dir, exist_ok=True)

    best_path = os.path.join(model_out_dir, "best.pt")
    onnx_path = os.path.join(model_out_dir, "best.onnx")
    meta_path = os.path.join(model_out_dir, "meta.json")
    curves_png_path = os.path.join(model_out_dir, "curves_loss_acc.png")

    test_pred_csv = os.path.join(model_out_dir, "test_predictions.csv")
    test_report_txt = os.path.join(model_out_dir, "test_report.txt")
    test_metrics_json = os.path.join(model_out_dir, "test_metrics.json")
    test_cm_npy = os.path.join(model_out_dir, "test_confusion_matrix.npy")
    test_cm_png = os.path.join(model_out_dir, "test_confusion_matrix.png")

    # Transforms
    train_tf, val_tf, mean, std = build_transforms(img_size)

    train_ds = datasets.ImageFolder(TRAIN_DIR, transform=train_tf)
    val_ds = datasets.ImageFolder(VAL_DIR, transform=val_tf)

    class_names = train_ds.classes
    num_classes = len(class_names)

    print(f"\n==============================")
    print(f"[MODEL] {model_name} | img_size={img_size}")
    print(f"[INFO] num_classes = {num_classes}")
    print("[INFO] classes =", class_names)

    class_weights, counts = compute_class_weights(train_ds)
    print("[INFO] class counts  =", counts.tolist())
    print("[INFO] class weights =", [round(x, 6) for x in class_weights.tolist()])

    train_loader = DataLoader(
        train_ds, batch_size=BATCH_SIZE, shuffle=True,
        num_workers=NUM_WORKERS, pin_memory=True
    )
    val_loader = DataLoader(
        val_ds, batch_size=BATCH_SIZE, shuffle=False,
        num_workers=NUM_WORKERS, pin_memory=True
    )

    # Build model
    model, arch = build_model(model_name, num_classes)
    model = model.to(device)

    criterion = nn.CrossEntropyLoss(weight=class_weights.to(device))
    optimizer = optim.AdamW(model.parameters(), lr=LR, weight_decay=WEIGHT_DECAY)
    scheduler = optim.lr_scheduler.ReduceLROnPlateau(optimizer, mode="min", factor=0.5, patience=2)

    use_amp = (device.type == "cuda")
    scaler = torch.amp.GradScaler('cuda', enabled=use_amp)

    best_val_acc = -1.0
    best_epoch = 0
    bad_epochs = 0

    train_losses, val_losses = [], []
    train_accs, val_accs = [], []

    epochs_ran = 0
    saved_at_least_once = False

    for epoch in range(1, EPOCHS + 1):
        epochs_ran = epoch
        model.train()
        total, correct = 0, 0
        running_loss = 0.0

        for x, y in tqdm(train_loader, desc=f"Train Epoch {epoch}", leave=False):
            x = x.to(device, non_blocking=True)
            y = y.to(device, non_blocking=True)
            optimizer.zero_grad(set_to_none=True)

            logits, aux = forward_logits(model, x, model_name, use_amp=use_amp)
            logits_fp32 = logits.float()

            if model_name == "inception_v3" and aux is not None and model.training:
                aux_fp32 = aux.float()
                loss_main = criterion(logits_fp32, y)
                loss_aux  = criterion(aux_fp32, y)
                loss = loss_main + 0.4 * loss_aux
            else:
                loss = criterion(logits_fp32, y)

            scaler.scale(loss).backward()
            scaler.step(optimizer)
            scaler.update()

            running_loss += loss.item() * x.size(0)
            pred = logits.argmax(1)
            correct += (pred == y).sum().item()
            total += y.size(0)

        train_loss = running_loss / max(total, 1)
        train_acc = correct / max(total, 1)

        val_loss, val_acc = evaluate(model, val_loader, device, criterion, model_name)
        scheduler.step(val_loss)

        train_losses.append(train_loss)
        val_losses.append(val_loss)
        train_accs.append(train_acc)
        val_accs.append(val_acc)

        lr_now = optimizer.param_groups[0]["lr"]
        print(f"[{model_name}][E{epoch:03d}] lr={lr_now:.2e} "
              f"train_loss={train_loss:.4f} train_acc={train_acc:.4f} | "
              f"val_loss={val_loss:.4f} val_acc={val_acc:.4f}")

        # ✅ val_acc 變好才存 best.pt
        if val_acc > best_val_acc:
            best_val_acc = val_acc
            best_epoch = epoch
            bad_epochs = 0

            torch.save({
                "model_state": model.state_dict(),
                "class_names": class_names,
                "img_size": img_size,
                "mean": mean,
                "std": std,
                "arch": arch,
                "best_val_acc": float(best_val_acc),
                "epoch": int(epoch),
            }, best_path)

            saved_at_least_once = True
            print(f"[SAVE] {model_name} best.pt -> {best_path} (val_acc={best_val_acc:.4f})")
        else:
            bad_epochs += 1
            if bad_epochs >= PATIENCE:
                print(f"[EARLY STOP] {model_name}: val_acc {PATIENCE} 次沒提升，停止。")
                break

    # 若完全沒存到 best.pt（極端情況），至少存最後一輪，避免後面 ONNX/測試都做不了
    if not saved_at_least_once:
        torch.save({
            "model_state": model.state_dict(),
            "class_names": class_names,
            "img_size": img_size,
            "mean": mean,
            "std": std,
            "arch": arch,
            "best_val_acc": float(best_val_acc),
            "epoch": int(epochs_ran),
        }, best_path)
        best_epoch = epochs_ran
        print(f"[SAVE-FALLBACK] {model_name} 沒有 best 提升，已存最後權重 -> {best_path}")

    # 1) 存曲線（固定顯示到 EPOCHS）
    plot_curves_fixed_epochs(train_losses, val_losses, train_accs, val_accs, curves_png_path, total_epochs=EPOCHS)
    print(f"[INFO] saved curves -> {curves_png_path}")

    # 2) 存 meta.json
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump({
            "arch": arch,
            "model_name": model_name,
            "train_dir": TRAIN_DIR,
            "val_dir": VAL_DIR,
            "test_dir": TEST_DIR,
            "out_dir": model_out_dir,
            "img_size": img_size,
            "epochs_setting": EPOCHS,
            "epochs_ran": int(epochs_ran),
            "batch_size": BATCH_SIZE,
            "lr": LR,
            "weight_decay": WEIGHT_DECAY,
            "best_val_acc": float(best_val_acc),
            "best_epoch": int(best_epoch),
            "classes": class_names,
            "curves_png": curves_png_path,
        }, f, ensure_ascii=False, indent=2)

    # 3) Test 評估 + 預測結果輸出
    metrics_row = {
        "model": model_name,
        "img_size": img_size,
        "epochs_setting": EPOCHS,
        "epochs_ran": int(epochs_ran),
        "best_val_acc": float(best_val_acc),
        "test_accuracy": None,
        "precision_weighted": None,
        "recall_weighted": None,
        "f1_weighted": None,
        "precision_macro": None,
        "recall_macro": None,
        "f1_macro": None,
        "pt_path": best_path,
        "onnx_path": onnx_path,
    }

    if os.path.isdir(TEST_DIR):
        print(f"[INFO] {model_name} Start testing.")

        test_ds = datasets.ImageFolder(TEST_DIR, transform=val_tf)
        test_loader = DataLoader(
            test_ds, batch_size=BATCH_SIZE, shuffle=False,
            num_workers=NUM_WORKERS, pin_memory=True
        )

        ckpt = torch.load(best_path, map_location=device)
        model.load_state_dict(ckpt["model_state"])
        model.eval()

        y_true, y_pred, confs, paths = predict_with_paths(model, test_ds, test_loader, device, model_name)

        acc = accuracy_score(y_true, y_pred)

        p_macro, r_macro, f1_macro, _ = precision_recall_fscore_support(
            y_true, y_pred, average="macro", zero_division=0
        )
        p_w, r_w, f1_w, _ = precision_recall_fscore_support(
            y_true, y_pred, average="weighted", zero_division=0
        )

        report = classification_report(
            y_true, y_pred,
            target_names=test_ds.classes,
            digits=4,
            zero_division=0
        )

        cm = confusion_matrix(y_true, y_pred)

        with open(test_report_txt, "w", encoding="utf-8") as f:
            f.write(report)

        with open(test_metrics_json, "w", encoding="utf-8") as f:
            json.dump({
                "accuracy": float(acc),
                "precision_macro": float(p_macro),
                "recall_macro": float(r_macro),
                "f1_macro": float(f1_macro),
                "precision_weighted": float(p_w),
                "recall_weighted": float(r_w),
                "f1_weighted": float(f1_w),
                "num_samples": int(len(y_true)),
            }, f, ensure_ascii=False, indent=2)

        np.save(test_cm_npy, cm)
        plot_confusion_matrix(cm, test_ds.classes, test_cm_png)

        with open(test_pred_csv, "w", encoding="utf-8") as f:
            f.write("path,true_label,true_name,pred_label,pred_name,confidence\n")
            for path, yt, yp, cf in zip(paths, y_true, y_pred, confs):
                true_name = test_ds.classes[yt]
                pred_name = test_ds.classes[yp]
                f.write(f"\"{path}\",{yt},\"{true_name}\",{yp},\"{pred_name}\",{cf:.6f}\n")

        print(f"[{model_name}][TEST] acc={acc:.4f} P(w)={p_w:.4f} R(w)={r_w:.4f} F1(w)={f1_w:.4f}")

        metrics_row.update({
            "test_accuracy": float(acc),
            "precision_weighted": float(p_w),
            "recall_weighted": float(r_w),
            "f1_weighted": float(f1_w),
            "precision_macro": float(p_macro),
            "recall_macro": float(r_macro),
            "f1_macro": float(f1_macro),
        })
    else:
        print(f"[WARN] TEST_DIR not found: {TEST_DIR}")

    # 匯出
    try:
        export_onnx_from_best_pt(
            best_pt_path=best_path,
            model_name=model_name,
            img_size=img_size,
            num_classes=num_classes,
            onnx_path=onnx_path
        )
        print(f"[INFO] saved onnx -> {onnx_path}")
    except Exception as e:
        print(f"[ERROR] ONNX export failed for {model_name}: {e}")

    return metrics_row


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"[INFO] device = {device}")

    all_rows = []
    for spec in MODEL_SPECS:
        row = run_one_model(spec["name"], spec["img_size"], device)
        all_rows.append(row)

    # 全部模型評估指標彙整輸出 Excel
    excel_path = os.path.join(OUT_DIR, "summary_metrics.xlsx")
    write_metrics_excel(all_rows, excel_path)
    print(f"[INFO] saved excel summary -> {excel_path}")


if __name__ == "__main__":
    main()
